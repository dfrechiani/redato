#!/usr/bin/env python3
"""Seed das 63 cartas estruturais do jogo "Redação em Jogo" (Fase 2).

Lê o arquivo `cartas_redacao_em_jogo.xlsx` (planilha "Estruturais
(lacuna)" — 3 colunas: Código, Tipo, Texto) e escreve em
`cartas_estruturais` no Postgres do portal.

As 63 estruturais são COMPARTILHADAS entre os 7 minidecks temáticos
(Saúde Mental, Inclusão Digital, etc.). Por isso não têm FK pra
`jogos_minideck` — são catalogadas isoladamente. Cartas temáticas
(P/R/K/A/AC/ME/F) ficam em `cartas_lacuna` e são populadas por outro
script (`seed_minideck.py`, Fase 5).

Modos:
- `--dry-run` (default): só relata o que seria feito.
- `--apply`: executa o UPSERT.

Idempotente: re-rodar `--apply` é seguro. Match por `codigo`. Se o
texto da carta mudar no xlsx (Daniel reescreveu), `--apply` atualiza.
Se nada mudou, é no-op (count=63 inserted_or_updated mas nenhum row
realmente alterado).

Validações:
- xlsx tem que existir
- Aba "Estruturais (lacuna)" tem que ter exatamente 63 rows com
  prefixo "E" (configurável via --expected-count)
- Toda secao tem que mapear pra um valor válido em `SECOES_ESTRUTURAIS`
- Toda cor é derivada da secao via `COR_POR_SECAO` (xlsx não tem
  coluna cor)
- Lacunas são extraídas do texto via regex `\\[([A-Z_]+)\\]`

Uso:
    cd <repo_root>
    python scripts/seed_cartas_estruturais.py             # dry-run
    python scripts/seed_cartas_estruturais.py --apply

Variáveis de ambiente:
    DATABASE_URL          Postgres do portal (obrigatório em --apply)
    SEED_CARTAS_XLSX      Override do path do xlsx (default
                          `backend/notamil-backend/data/seeds/
                          cartas_redacao_em_jogo.xlsx` no repo).
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


REPO = Path(__file__).resolve().parent.parent
BACKEND = REPO / "backend" / "notamil-backend"
sys.path.insert(0, str(BACKEND))

try:
    from dotenv import load_dotenv  # type: ignore[import-untyped]
    load_dotenv(BACKEND / ".env")
except ImportError:
    pass


logger = logging.getLogger("seed_cartas_estruturais")


# ──────────────────────────────────────────────────────────────────────
# Constantes
# ──────────────────────────────────────────────────────────────────────

# Default xlsx path (tem que ser commit no repo). Override via
# SEED_CARTAS_XLSX env ou --file.
DEFAULT_XLSX = (
    BACKEND / "data" / "seeds" / "cartas_redacao_em_jogo.xlsx"
)
SHEET_NAME = "Estruturais (lacuna)"
EXPECTED_COUNT_DEFAULT = 63

# Mapping "Tipo" do xlsx → enum `secao` da tabela. xlsx usa label
# humano com acento; DB usa identifier ASCII pra evitar pegadinha
# de codepage e simplificar CHECK constraint.
SECAO_BY_TIPO_XLSX = {
    "ABERTURA": "ABERTURA",
    "TESE": "TESE",
    "TÓPICO Dev1": "TOPICO_DEV1",
    "ARGUMENTO Dev1": "ARGUMENTO_DEV1",
    "REPERTÓRIO Dev1": "REPERTORIO_DEV1",
    "TÓPICO Dev2": "TOPICO_DEV2",
    "ARGUMENTO Dev2": "ARGUMENTO_DEV2",
    "REPERTÓRIO Dev2": "REPERTORIO_DEV2",
    "RETOMADA": "RETOMADA",
    "PROPOSTA": "PROPOSTA",
}

# Mapping autoritativo. Vem do README do xlsx ("Cores estruturais:
# Azul = Introdução, Amarelo = Dev1, Verde = Dev2, Laranja = Conclusão").
# Replicado no models.COR_POR_SECAO pra ficar acessível como import.
COR_POR_SECAO = {
    "ABERTURA": "AZUL", "TESE": "AZUL",
    "TOPICO_DEV1": "AMARELO", "ARGUMENTO_DEV1": "AMARELO",
    "REPERTORIO_DEV1": "AMARELO",
    "TOPICO_DEV2": "VERDE", "ARGUMENTO_DEV2": "VERDE",
    "REPERTORIO_DEV2": "VERDE",
    "RETOMADA": "LARANJA", "PROPOSTA": "LARANJA",
}

# Regex pra extrair placeholders [PROBLEMA], [REPERTORIO], etc. Captura
# em ordem de aparição. Chars válidos: A-Z + underscore — barra
# match acidental com "[Lei 13.146]" ou outras citações que não são
# placeholders do jogo.
_PLACEHOLDER_RE = re.compile(r"\[([A-Z_]+)\]")
PLACEHOLDERS_VALIDOS = frozenset({
    "PROBLEMA", "REPERTORIO", "PALAVRA_CHAVE", "AGENTE", "ACAO_MEIO",
})


# ──────────────────────────────────────────────────────────────────────
# Setup
# ──────────────────────────────────────────────────────────────────────

def _setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] %(levelname)s %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )


def _resolve_xlsx_path(override: Optional[str]) -> Path:
    """Ordem de prioridade: --file → SEED_CARTAS_XLSX env → default."""
    if override:
        return Path(override)
    env_override = os.getenv("SEED_CARTAS_XLSX")
    if env_override:
        return Path(env_override)
    return DEFAULT_XLSX


# ──────────────────────────────────────────────────────────────────────
# Parsing
# ──────────────────────────────────────────────────────────────────────

def _extract_lacunas(texto: str) -> List[str]:
    """Lista placeholders presentes em `texto`, em ordem de aparição.
    Sem deduplicação — se [PALAVRA_CHAVE] aparece 2x, retorna 2x. Útil
    pro frontend renderizar 2 slots distintos."""
    return [
        m.group(1) for m in _PLACEHOLDER_RE.finditer(texto)
        if m.group(1) in PLACEHOLDERS_VALIDOS
    ]


def parse_xlsx(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Lê a aba "Estruturais (lacuna)" e retorna lista de dicts:
        {codigo, secao, cor, texto, lacunas, ordem}

    `ordem` = índice da row no xlsx (1, 2, 3, ...). Define ordem
    canônica de exibição na UI quando agrupado por seção (a planilha
    foi montada por uma corretora-parceira nesse arrangement).
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"xlsx não encontrado em {xlsx_path}. Defina --file ou "
            f"SEED_CARTAS_XLSX env var.",
        )

    # Import lazy pra script poder rodar `--help` sem ter openpyxl.
    import openpyxl  # type: ignore[import-untyped]

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(
            f"Aba {SHEET_NAME!r} não encontrada no xlsx. "
            f"Abas disponíveis: {wb.sheetnames}",
        )
    ws = wb[SHEET_NAME]

    rows: List[Dict[str, Any]] = []
    ordem = 0
    tipos_nao_mapeados: set = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        if not row or row[0] is None:
            continue  # row vazia (possível trailing)
        codigo, tipo_xlsx, texto = row[0], row[1], row[2]
        codigo = str(codigo).strip()
        # Filtro defensivo: cartas estruturais sempre começam com "E".
        # Se aparecer outro prefixo, ignora silencioso (xlsx pode ter
        # outras seções no futuro que não queremos seedar aqui).
        if not re.match(r"^E\d+$", codigo):
            continue
        if not tipo_xlsx or texto is None:
            raise RuntimeError(
                f"Row {i} (codigo={codigo}) tem campo vazio: "
                f"tipo={tipo_xlsx!r} texto={texto!r}",
            )
        tipo_str = str(tipo_xlsx).strip()
        secao = SECAO_BY_TIPO_XLSX.get(tipo_str)
        if secao is None:
            tipos_nao_mapeados.add(tipo_str)
            continue
        cor = COR_POR_SECAO[secao]
        texto_str = str(texto).strip()
        lacunas = _extract_lacunas(texto_str)
        ordem += 1
        rows.append({
            "codigo": codigo,
            "secao": secao,
            "cor": cor,
            "texto": texto_str,
            "lacunas": lacunas,
            "ordem": ordem,
        })

    if tipos_nao_mapeados:
        raise RuntimeError(
            f"Tipos no xlsx sem mapeamento: {sorted(tipos_nao_mapeados)}. "
            f"Adicione em SECAO_BY_TIPO_XLSX ou corrija o xlsx.",
        )

    return rows


# ──────────────────────────────────────────────────────────────────────
# DB — UPSERT idempotente
# ──────────────────────────────────────────────────────────────────────

def upsert(
    rows: List[Dict[str, Any]],
    *,
    apply: bool,
) -> Tuple[int, int]:
    """UPSERT em `cartas_estruturais` por `codigo`. Retorna
    (inserted, updated). Em dry-run os valores refletem o que seria
    feito (queries de SELECT pra contar diff)."""
    from sqlalchemy import select
    from sqlalchemy.dialects.postgresql import insert as pg_insert
    from sqlalchemy.orm import Session

    from redato_backend.portal.db import get_engine
    from redato_backend.portal.models import CartaEstrutural

    engine = get_engine()
    inserted = 0
    updated = 0

    with Session(engine) as session:
        # Snapshot atual (codigo → row) pra calcular diff sem precisar
        # do RETURNING do UPSERT.
        existentes = {
            r.codigo: r for r in session.execute(
                select(CartaEstrutural)
            ).scalars()
        }

        for row in rows:
            cur = existentes.get(row["codigo"])
            if cur is None:
                inserted += 1
            else:
                # Se algum campo mudou, conta como update.
                if (
                    cur.secao != row["secao"]
                    or cur.cor != row["cor"]
                    or cur.texto != row["texto"]
                    or list(cur.lacunas or []) != row["lacunas"]
                    or cur.ordem != row["ordem"]
                ):
                    updated += 1

        if apply:
            # ON CONFLICT (codigo) DO UPDATE SET ... — atomic e idempotente.
            stmt = pg_insert(CartaEstrutural.__table__).values(rows)
            stmt = stmt.on_conflict_do_update(
                index_elements=["codigo"],
                set_={
                    "secao": stmt.excluded.secao,
                    "cor": stmt.excluded.cor,
                    "texto": stmt.excluded.texto,
                    "lacunas": stmt.excluded.lacunas,
                    "ordem": stmt.excluded.ordem,
                },
            )
            session.execute(stmt)
            session.commit()
        else:
            session.rollback()

    return inserted, updated


# ──────────────────────────────────────────────────────────────────────
# Resumo / logging
# ──────────────────────────────────────────────────────────────────────

def _resumir_por_secao(rows: List[Dict[str, Any]]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for r in rows:
        out[r["secao"]] = out.get(r["secao"], 0) + 1
    return out


def _print_resumo(rows: List[Dict[str, Any]]) -> None:
    by_secao = _resumir_por_secao(rows)
    logger.info("Resumo por seção:")
    for secao, n in sorted(by_secao.items()):
        cor = COR_POR_SECAO.get(secao, "?")
        logger.info("  %-18s (%-7s) → %d cartas", secao, cor, n)
    logger.info("  TOTAL = %d", len(rows))


# ──────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────

def main() -> int:
    _setup_logging()

    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--apply", action="store_true",
        help="Executa o UPSERT em Postgres. Sem isso, é dry-run.",
    )
    parser.add_argument(
        "--file", type=str, default=None,
        help=(
            "Override do path do xlsx. Default: "
            "backend/notamil-backend/data/seeds/cartas_redacao_em_jogo.xlsx"
        ),
    )
    parser.add_argument(
        "--expected-count", type=int, default=EXPECTED_COUNT_DEFAULT,
        help=f"Contagem esperada (default: {EXPECTED_COUNT_DEFAULT}).",
    )
    args = parser.parse_args()

    xlsx_path = _resolve_xlsx_path(args.file)
    logger.info(
        "Modo=%s — lendo %s",
        "APPLY" if args.apply else "DRY-RUN", xlsx_path,
    )

    try:
        rows = parse_xlsx(xlsx_path)
    except FileNotFoundError as exc:
        logger.error(str(exc))
        return 2
    except RuntimeError as exc:
        logger.error("Falha ao parsear xlsx: %s", exc)
        return 3

    _print_resumo(rows)

    if len(rows) != args.expected_count:
        logger.error(
            "Contagem fora do esperado: rows=%d expected=%d. "
            "Reveja o xlsx ou ajuste --expected-count.",
            len(rows), args.expected_count,
        )
        return 4

    try:
        inserted, updated = upsert(rows, apply=args.apply)
    except Exception:
        logger.exception("Falha no UPSERT")
        return 5

    if args.apply:
        logger.info(
            "✓ APPLY ok — inserted=%d updated=%d unchanged=%d total=%d",
            inserted, updated, len(rows) - inserted - updated, len(rows),
        )
    else:
        logger.info(
            "DRY-RUN — would insert=%d update=%d (no DB writes). "
            "Use --apply pra commitar.",
            inserted, updated,
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
