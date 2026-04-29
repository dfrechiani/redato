#!/usr/bin/env python3
"""Seed de minideck temático do jogo "Redação em Jogo" (Fase 2 passo 2).

Lê uma aba do `cartas_redacao_em_jogo.xlsx` correspondente a um tema
(Saúde Mental, Inclusão Digital, etc.) e popula:

1. Uma row em `jogos_minideck` (catálogo do tema)
2. ~104 rows em `cartas_lacuna` (cartas P/R/K/A/AC/ME/F do tema)

As 63 `cartas_estruturais` (E01-E63) já são populadas pelo passo 1
(`seed_cartas_estruturais.py`) e são compartilhadas entre todos os
minidecks — esse script NÃO toca nelas.

Modos:

    python scripts/seed_minideck.py <tema>             # dry-run de 1 tema
    python scripts/seed_minideck.py <tema> --apply     # commita 1 tema
    python scripts/seed_minideck.py --all              # dry-run de todos
    python scripts/seed_minideck.py --all --apply      # commita todos
    python scripts/seed_minideck.py --list             # lista temas

`<tema>` é o slug snake_case (saude_mental, inclusao_digital, ...).
Use `--list` pra ver todos.

Idempotência:
- `jogos_minideck`: ON CONFLICT (tema) DO UPDATE — atualiza
  nome_humano, descricao, updated_at.
- `cartas_lacuna`: ON CONFLICT (minideck_id, codigo) DO UPDATE —
  atualiza tipo + conteudo. Permite editar cartas no xlsx e
  re-seedar sem duplicar.

Cada tema roda em sua própria transação. Em `--all`, falha de um
tema NÃO bloqueia os outros — script continua e relata stats por
tema no final. Mas dentro de UM tema, falha em qualquer carta gera
rollback completo daquele tema (atomicidade por minideck).

Validações antes de gravar:
- Aba existe no xlsx
- Tem pelo menos 1 carta de cada tipo obrigatório (P, R, K, A, AC,
  ME, F)
- Total >= 50 cartas (sanity — minideck real tem ~104)
- Todo `codigo` bate com `prefixo_do_tipo + 2 dígitos zero-padded`
- Nenhum `codigo` começa com `E` (reservado pra cartas_estruturais)

Variáveis de ambiente:
    DATABASE_URL          Postgres (obrigatório em --apply)
    SEED_CARTAS_XLSX      Override do path do xlsx (default
                          backend/notamil-backend/data/seeds/
                          cartas_redacao_em_jogo.xlsx)
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


REPO = Path(__file__).resolve().parent.parent
BACKEND = REPO / "backend" / "notamil-backend"
sys.path.insert(0, str(BACKEND))

try:
    from dotenv import load_dotenv  # type: ignore[import-untyped]
    load_dotenv(BACKEND / ".env")
except ImportError:
    pass


logger = logging.getLogger("seed_minideck")


# ──────────────────────────────────────────────────────────────────────
# Constantes
# ──────────────────────────────────────────────────────────────────────

DEFAULT_XLSX = (
    BACKEND / "data" / "seeds" / "cartas_redacao_em_jogo.xlsx"
)

# Mapping aba xlsx → slug (tema canônico no DB). slug é o que vai pra
# URL/path-param/JSON e fica imutável; nome_humano (label da aba) pode
# ser editado sem quebrar referências.
TEMAS_MAPEAMENTO: Dict[str, str] = {
    "Saúde Mental": "saude_mental",
    "Inclusão Digital": "inclusao_digital",
    "Violência contra a Mulher": "violencia_contra_mulher",
    "Educação Financeira": "educacao_financeira",
    "Gênero e Diversidade": "genero_diversidade",
    "Meio Ambiente": "meio_ambiente",
    "Família e Sociedade": "familia_sociedade",
}
# Inverso pra `<tema>` na CLI virar nome de aba.
ABA_BY_TEMA: Dict[str, str] = {v: k for k, v in TEMAS_MAPEAMENTO.items()}

# Mapping tipo no xlsx (label humano com acento) → enum DB (ASCII).
# Aceita também aliases curtos (P/R/K/A/AC/ME/F) e variantes ASCII
# pra defensividade.
TIPO_BY_LABEL: Dict[str, str] = {
    # Labels do xlsx (forma canônica que aparece na planilha)
    "PROBLEMA": "PROBLEMA",
    "REPERTÓRIO": "REPERTORIO",
    "PALAVRA-CHAVE": "PALAVRA_CHAVE",
    "AGENTE": "AGENTE",
    "AÇÃO": "ACAO",
    "MEIO": "MEIO",
    "FIM": "FIM",
    # Variantes ASCII (defensivo — Daniel pode editar removendo acento)
    "REPERTORIO": "REPERTORIO",
    "PALAVRA_CHAVE": "PALAVRA_CHAVE",
    "PALAVRACHAVE": "PALAVRA_CHAVE",
    "ACAO": "ACAO",
    # Aliases curtos do briefing (P/R/K/A/AC/ME/F)
    "P": "PROBLEMA",
    "R": "REPERTORIO",
    "K": "PALAVRA_CHAVE",
    "A": "AGENTE",
    "AC": "ACAO",
    "ME": "MEIO",
    "F": "FIM",
}

# Prefixo do `codigo` esperado por tipo. ME e AC são 2 letras pra
# diferenciar de M/A simples. Codigo final = prefixo + 2 dígitos.
PREFIXO_BY_TIPO: Dict[str, str] = {
    "PROBLEMA": "P",
    "REPERTORIO": "R",
    "PALAVRA_CHAVE": "K",
    "AGENTE": "A",
    "ACAO": "AC",
    "MEIO": "ME",
    "FIM": "F",
}

# Tipos que TODO minideck precisa ter (mínimo 1 carta cada). Sem isso
# o jogo trava — por exemplo, se minideck não tem AGENTE, o aluno
# escolheu PROPOSTA E51 com [AGENTE] e não tem carta pra preencher.
TIPOS_OBRIGATORIOS = (
    "PROBLEMA", "REPERTORIO", "PALAVRA_CHAVE",
    "AGENTE", "ACAO", "MEIO", "FIM",
)
# Sanity — minideck normal tem ~104 cartas. < 50 é placeholder/teste.
MIN_CARTAS_TOTAL = 50

# Default de série pro minideck (decisão pedagógica G.1.5: minidecks
# completos são pra 2ª série EM). `serie` é só campo descritivo, não
# afeta lógica do jogo.
DEFAULT_SERIE = "2S"


# ──────────────────────────────────────────────────────────────────────
# Setup / helpers
# ──────────────────────────────────────────────────────────────────────

def _setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] %(levelname)s %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )


def _resolve_xlsx_path(override: Optional[str]) -> Path:
    if override:
        return Path(override)
    env_override = os.getenv("SEED_CARTAS_XLSX")
    if env_override:
        return Path(env_override)
    return DEFAULT_XLSX


@dataclass
class CartaParseError:
    """Erro de parse de uma row do xlsx — coletado pra reportar tudo
    de uma vez (não para no primeiro erro)."""
    excel_row: int
    codigo: Optional[str]
    motivo: str


# ──────────────────────────────────────────────────────────────────────
# Parser
# ──────────────────────────────────────────────────────────────────────

def _normaliza_tipo_label(raw: Any) -> Optional[str]:
    """Trim + uppercase. Aceita None / non-str → None (caller trata
    como erro)."""
    if raw is None:
        return None
    s = str(raw).strip().upper()
    return s or None


def _valida_codigo(codigo: str, tipo_db: str) -> Optional[str]:
    """Retorna mensagem de erro ou None se OK. Regras:
    - Prefixo bate com PREFIXO_BY_TIPO[tipo_db]
    - Restante = exatamente 2 dígitos
    - Não começa com 'E' (reservado pra estruturais)
    """
    if codigo.startswith("E"):
        return f"código {codigo!r} começa com 'E' (reservado pra cartas_estruturais)"
    prefix = PREFIXO_BY_TIPO[tipo_db]
    pat = rf"^{re.escape(prefix)}\d{{2}}$"
    if not re.match(pat, codigo):
        return (
            f"código {codigo!r} não bate com padrão "
            f"{prefix}NN (2 dígitos zero-padded) esperado pra tipo "
            f"{tipo_db}"
        )
    return None


def parse_aba(
    xlsx_path: Path, aba_nome: str,
) -> Tuple[List[Dict[str, Any]], List[CartaParseError]]:
    """Lê uma aba e retorna (rows_validas, erros). Caller decide se
    aborta no primeiro erro — função NÃO levanta exceção em row
    inválida, coleta tudo pra log."""
    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"xlsx não encontrado em {xlsx_path}. Defina --file ou "
            f"SEED_CARTAS_XLSX env var.",
        )

    import openpyxl  # type: ignore[import-untyped]

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if aba_nome not in wb.sheetnames:
        raise RuntimeError(
            f"Aba {aba_nome!r} não encontrada no xlsx. Abas "
            f"disponíveis: {wb.sheetnames}",
        )
    ws = wb[aba_nome]

    rows: List[Dict[str, Any]] = []
    erros: List[CartaParseError] = []

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue  # header
        if not row or all(c is None for c in row):
            continue  # row vazia
        codigo_raw = row[0]
        tipo_raw = row[1] if len(row) > 1 else None
        conteudo_raw = row[2] if len(row) > 2 else None

        codigo = (str(codigo_raw).strip()
                  if codigo_raw is not None else "")
        tipo_label = _normaliza_tipo_label(tipo_raw)
        conteudo = (str(conteudo_raw).strip()
                    if conteudo_raw is not None else "")

        # Skip + warn se row tá parcial — Daniel pode ter dragged uma
        # célula e deixado um codigo solo. Vira erro pra ele revisar.
        if not codigo:
            # Row sem codigo + sem outros campos = trailing vazia,
            # já filtrado acima. Aqui é codigo ausente mas com tipo
            # ou conteúdo — anomalia de fato.
            if tipo_label or conteudo:
                erros.append(CartaParseError(
                    excel_row=i, codigo=None,
                    motivo=("row sem codigo mas com tipo/conteúdo "
                            "preenchido"),
                ))
            continue
        if tipo_label is None:
            erros.append(CartaParseError(
                excel_row=i, codigo=codigo,
                motivo="tipo vazio",
            ))
            continue
        if not conteudo:
            erros.append(CartaParseError(
                excel_row=i, codigo=codigo,
                motivo="conteudo vazio",
            ))
            continue

        tipo_db = TIPO_BY_LABEL.get(tipo_label)
        if tipo_db is None:
            erros.append(CartaParseError(
                excel_row=i, codigo=codigo,
                motivo=(
                    f"tipo {tipo_label!r} não reconhecido. Aceitos: "
                    f"{sorted(set(TIPO_BY_LABEL.keys()))}"
                ),
            ))
            continue

        err = _valida_codigo(codigo, tipo_db)
        if err:
            erros.append(CartaParseError(
                excel_row=i, codigo=codigo, motivo=err,
            ))
            continue

        rows.append({
            "codigo": codigo,
            "tipo": tipo_db,
            "conteudo": conteudo,
            "_excel_row": i,
        })

    return rows, erros


# ──────────────────────────────────────────────────────────────────────
# Validação semântica (post-parse, pre-write)
# ──────────────────────────────────────────────────────────────────────

def _resumir_por_tipo(rows: List[Dict[str, Any]]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for r in rows:
        out[r["tipo"]] = out.get(r["tipo"], 0) + 1
    return out


def valida_minideck(
    rows: List[Dict[str, Any]], tema: str,
) -> List[str]:
    """Valida regras de negócio do minideck (post-parse). Retorna
    lista de erros — vazia = OK."""
    problemas: List[str] = []

    if len(rows) < MIN_CARTAS_TOTAL:
        problemas.append(
            f"total {len(rows)} cartas < {MIN_CARTAS_TOTAL} mínimo. "
            f"Minideck normal tem ~104 — verifica se aba {tema!r} "
            f"foi lida completa.",
        )

    por_tipo = _resumir_por_tipo(rows)
    for tipo in TIPOS_OBRIGATORIOS:
        if por_tipo.get(tipo, 0) == 0:
            problemas.append(
                f"tipo {tipo} obrigatório ausente. Sem ele o jogo "
                f"trava (placeholder do estrutural não preenche).",
            )

    # Codigos duplicados dentro da mesma aba — UPSERT no DB resolveria
    # silenciosamente (último ganha), mas isso indica erro do xlsx.
    codigos = [r["codigo"] for r in rows]
    if len(codigos) != len(set(codigos)):
        from collections import Counter
        dup = [c for c, n in Counter(codigos).items() if n > 1]
        problemas.append(
            f"códigos duplicados na aba: {dup}",
        )

    return problemas


# ──────────────────────────────────────────────────────────────────────
# DB — UPSERT atômico por tema
# ──────────────────────────────────────────────────────────────────────

@dataclass
class TemaStats:
    tema: str
    minideck_inserted: bool = False
    minideck_updated: bool = False
    cartas_inserted: int = 0
    cartas_updated: int = 0
    cartas_unchanged: int = 0
    erro: Optional[str] = None


def upsert_tema(
    *, tema: str, nome_humano: str, rows: List[Dict[str, Any]],
    apply: bool, descricao: Optional[str] = None,
) -> TemaStats:
    """Upsert atômico do tema + cartas. Em apply=False roda os SELECTs
    pra contar diff e faz rollback. Em apply=True comita.

    Erro em qualquer carta dispara rollback do tema inteiro (zero
    estado parcial)."""
    from sqlalchemy import select
    from sqlalchemy.dialects.postgresql import insert as pg_insert
    from sqlalchemy.orm import Session

    from redato_backend.portal.db import get_engine
    from redato_backend.portal.models import CartaLacuna, JogoMinideck

    stats = TemaStats(tema=tema)
    engine = get_engine()

    try:
        with Session(engine) as session:
            # ──────────────────────────────────────────────────────
            # 1. Minideck — UPSERT por `tema`
            # ──────────────────────────────────────────────────────
            existente_md = session.execute(
                select(JogoMinideck).where(JogoMinideck.tema == tema)
            ).scalar_one_or_none()

            md_payload: Dict[str, Any] = {
                "tema": tema,
                "nome_humano": nome_humano,
                "serie": DEFAULT_SERIE,
                "ativo": True,
            }
            if descricao is not None:
                md_payload["descricao"] = descricao

            if existente_md is None:
                stats.minideck_inserted = True
            else:
                # Detectar update — relevante se nome_humano ou
                # descricao mudou no xlsx.
                if (existente_md.nome_humano != nome_humano
                        or (descricao is not None
                            and existente_md.descricao != descricao)):
                    stats.minideck_updated = True

            if apply:
                stmt = pg_insert(JogoMinideck.__table__).values([md_payload])
                set_ = {
                    "nome_humano": stmt.excluded.nome_humano,
                    "updated_at": stmt.excluded.created_at,  # touch
                }
                if descricao is not None:
                    set_["descricao"] = stmt.excluded.descricao
                stmt = stmt.on_conflict_do_update(
                    index_elements=["tema"], set_=set_,
                )
                session.execute(stmt)
                session.flush()
                # Re-fetch pra pegar o id (novo ou existente)
                existente_md = session.execute(
                    select(JogoMinideck).where(JogoMinideck.tema == tema)
                ).scalar_one()

            # Em dry-run sem md existente, não temos id pra inserir
            # cartas — só relatamos os counts e simulamos.
            md_id = (existente_md.id if existente_md else None)

            # ──────────────────────────────────────────────────────
            # 2. Cartas — UPSERT por (minideck_id, codigo)
            # ──────────────────────────────────────────────────────
            existentes_lacuna = {}
            if md_id is not None:
                existentes_lacuna = {
                    c.codigo: c for c in session.execute(
                        select(CartaLacuna).where(
                            CartaLacuna.minideck_id == md_id,
                        )
                    ).scalars()
                }

            for r in rows:
                cur = existentes_lacuna.get(r["codigo"])
                if cur is None:
                    stats.cartas_inserted += 1
                else:
                    if (cur.tipo != r["tipo"]
                            or cur.conteudo != r["conteudo"]):
                        stats.cartas_updated += 1
                    else:
                        stats.cartas_unchanged += 1

            if apply and md_id is not None:
                payload_cartas = [
                    {
                        "minideck_id": md_id,
                        "tipo": r["tipo"],
                        "codigo": r["codigo"],
                        "conteudo": r["conteudo"],
                    }
                    for r in rows
                ]
                if payload_cartas:
                    stmt = pg_insert(CartaLacuna.__table__).values(
                        payload_cartas,
                    )
                    stmt = stmt.on_conflict_do_update(
                        index_elements=["minideck_id", "codigo"],
                        set_={
                            "tipo": stmt.excluded.tipo,
                            "conteudo": stmt.excluded.conteudo,
                        },
                    )
                    session.execute(stmt)

            if apply:
                session.commit()
            else:
                session.rollback()
    except Exception as exc:  # noqa: BLE001
        stats.erro = repr(exc)
        logger.exception("Falha no upsert do tema %r", tema)

    return stats


# ──────────────────────────────────────────────────────────────────────
# Pipeline por tema (parse + valida + upsert)
# ──────────────────────────────────────────────────────────────────────

def processar_tema(
    *, tema: str, xlsx_path: Path, apply: bool,
) -> TemaStats:
    """Pipeline completo de 1 tema. Retorna stats com erro setado se
    qualquer etapa falhou — caller decide se aborta (--apply 1 tema)
    ou continua (--all)."""
    aba = ABA_BY_TEMA.get(tema)
    if aba is None:
        s = TemaStats(tema=tema)
        s.erro = (
            f"tema {tema!r} não está em TEMAS_MAPEAMENTO. "
            f"Use --list pra ver os disponíveis."
        )
        logger.error(s.erro)
        return s

    logger.info("══ Tema: %s (aba %r) ══", tema, aba)

    try:
        rows, erros = parse_aba(xlsx_path, aba)
    except (FileNotFoundError, RuntimeError) as exc:
        s = TemaStats(tema=tema)
        s.erro = str(exc)
        logger.error(s.erro)
        return s

    if erros:
        for e in erros:
            logger.error(
                "  Excel row %d codigo=%s: %s",
                e.excel_row, e.codigo, e.motivo,
            )
        s = TemaStats(tema=tema)
        s.erro = f"{len(erros)} erros de parse — abortando tema"
        return s

    # Resumo
    por_tipo = _resumir_por_tipo(rows)
    for tipo in sorted(por_tipo):
        prefix = PREFIXO_BY_TIPO.get(tipo, "?")
        logger.info(
            "  %-15s (%-2s) → %d cartas",
            tipo, prefix, por_tipo[tipo],
        )
    logger.info("  TOTAL = %d", len(rows))

    problemas = valida_minideck(rows, tema)
    if problemas:
        for p in problemas:
            logger.error("  ✗ %s", p)
        s = TemaStats(tema=tema)
        s.erro = f"validação semântica falhou ({len(problemas)} erros)"
        return s

    stats = upsert_tema(
        tema=tema, nome_humano=aba, rows=rows, apply=apply,
    )
    return stats


# ──────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────

def _print_stats(stats: TemaStats, apply: bool) -> None:
    if stats.erro:
        logger.error("✗ Tema %s falhou: %s", stats.tema, stats.erro)
        return
    label = "✓ APPLY" if apply else "DRY-RUN"
    md_status = (
        "minideck=NEW" if stats.minideck_inserted
        else ("minideck=UPDATED" if stats.minideck_updated
              else "minideck=unchanged")
    )
    if apply:
        logger.info(
            "%s %s — %s cartas: inserted=%d updated=%d unchanged=%d",
            label, stats.tema, md_status,
            stats.cartas_inserted, stats.cartas_updated,
            stats.cartas_unchanged,
        )
    else:
        logger.info(
            "%s %s — %s cartas: would insert=%d update=%d "
            "unchanged=%d",
            label, stats.tema, md_status,
            stats.cartas_inserted, stats.cartas_updated,
            stats.cartas_unchanged,
        )


def main() -> int:
    _setup_logging()
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "tema", nargs="?", default=None,
        help=(
            "Slug do tema (saude_mental, inclusao_digital, ...). "
            "Use --list pra ver todos."
        ),
    )
    parser.add_argument(
        "--apply", action="store_true",
        help="Comita no Postgres. Sem isso, é dry-run.",
    )
    parser.add_argument(
        "--all", action="store_true",
        help=("Processa todos os 7 temas. Falha de um NÃO bloqueia "
              "os outros — relata erros no resumo final."),
    )
    parser.add_argument(
        "--list", action="store_true",
        help="Lista temas disponíveis e sai.",
    )
    parser.add_argument(
        "--file", type=str, default=None,
        help="Override do xlsx path. Default: data/seeds/cartas_redacao_em_jogo.xlsx",
    )
    args = parser.parse_args()

    if args.list:
        print("Temas disponíveis:")
        for nome_humano, slug in TEMAS_MAPEAMENTO.items():
            print(f"  {slug:30s} ← aba {nome_humano!r}")
        return 0

    if args.all and args.tema:
        logger.error("Use --all OU <tema>, não ambos.")
        return 1
    if not args.all and not args.tema:
        logger.error(
            "Falta argumento. Uso: <tema> ou --all ou --list. "
            "Veja --help.",
        )
        return 1

    xlsx_path = _resolve_xlsx_path(args.file)
    logger.info(
        "Modo=%s — xlsx=%s",
        "APPLY" if args.apply else "DRY-RUN", xlsx_path,
    )

    if args.all:
        all_stats: List[TemaStats] = []
        for tema in TEMAS_MAPEAMENTO.values():
            s = processar_tema(
                tema=tema, xlsx_path=xlsx_path, apply=args.apply,
            )
            _print_stats(s, args.apply)
            all_stats.append(s)
        # Resumo final
        n_ok = sum(1 for s in all_stats if not s.erro)
        n_err = sum(1 for s in all_stats if s.erro)
        logger.info("══ Resumo ══")
        logger.info("  %d temas OK · %d com erro", n_ok, n_err)
        if n_err > 0:
            for s in all_stats:
                if s.erro:
                    logger.info("    ✗ %s: %s", s.tema, s.erro)
            return 6
        return 0

    # Single-tema
    if args.tema not in ABA_BY_TEMA:
        logger.error(
            "tema %r desconhecido. Disponíveis: %s",
            args.tema, list(ABA_BY_TEMA.keys()),
        )
        return 2
    s = processar_tema(
        tema=args.tema, xlsx_path=xlsx_path, apply=args.apply,
    )
    _print_stats(s, args.apply)
    return 0 if not s.erro else 6


if __name__ == "__main__":
    sys.exit(main())
