"""Importador de planilha SEDUC — parser, validador, sync ORM (M2).

Formato esperado (8 colunas, headers obrigatórios na 1ª linha):

    escola_id          | escola_nome
    coordenador_email  | coordenador_nome
    professor_email    | professor_nome
    turma_codigo       | turma_serie  (1S/2S/3S)

Uma linha = uma turma. Múltiplas turmas da mesma escola repetem
escola_id/escola_nome. Idempotência por chave natural:
- Escola por `codigo`
- Coordenador / Professor por `email`
- Turma por `(escola_id, codigo, ano_letivo)`

Spec: docs/redato/v3/REPORT_caminho2_realuse.md (seção 5).
"""
from __future__ import annotations

import csv
import os
import re
import unicodedata
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from email_validator import EmailNotValidError, validate_email
from sqlalchemy import select
from sqlalchemy.orm import Session

from redato_backend.portal.models import (
    AlunoTurma, Atividade, Coordenador, Escola, Professor, Turma,
)


# Headers obrigatórios (8 colunas) + opcionais (2 extras de M3).
REQUIRED_HEADERS = (
    "escola_id", "escola_nome",
    "coordenador_email", "coordenador_nome",
    "professor_email", "professor_nome",
    "turma_codigo", "turma_serie",
)
OPTIONAL_HEADERS = ("escola_estado", "escola_municipio")
EXPECTED_HEADERS = REQUIRED_HEADERS  # backcompat alias (testes)

VALID_SERIES = ("1S", "2S", "3S")

# Regex default pra escola_id (configurável via env). Aceita SEDUC-CE-001,
# CE-PUB-1234, etc. Letras maiúsculas + hífen + dígitos no fim.
DEFAULT_ESCOLA_ID_REGEX = r"^[A-Z]+-[A-Z]{2}-\d{3,}$"
TURMA_CODIGO_REGEX = re.compile(r"^[A-Za-z0-9]{1,8}$")


# ──────────────────────────────────────────────────────────────────────
# Estruturas
# ──────────────────────────────────────────────────────────────────────

@dataclass
class Issue:
    """Erro ou warning detectado durante validação ou sync."""
    line: Optional[int]   # 1-indexed (linha da planilha excluindo header)
    field: Optional[str]
    code: str             # código curto pra agrupar/filtrar
    message: str
    severity: str = "error"  # "error" | "warning"

    def to_dict(self) -> Dict[str, Any]:
        return {
            "line": self.line, "field": self.field, "code": self.code,
            "message": self.message, "severity": self.severity,
        }


@dataclass
class ImportReport:
    """Relatório consolidado de uma rodada de import."""
    modo: str  # "dry-run" | "commit"
    arquivo: Optional[str] = None
    linhas_lidas: int = 0
    escolas_novas: int = 0
    escolas_atualizadas: int = 0
    coordenadores_novos: int = 0
    coordenadores_atualizados: int = 0
    professores_novos: int = 0
    professores_atualizados: int = 0
    turmas_novas: int = 0
    turmas_atualizadas: int = 0
    warnings: List[Dict[str, Any]] = field(default_factory=list)
    erros: List[Dict[str, Any]] = field(default_factory=list)

    def add_issue(self, issue: Issue) -> None:
        bucket = self.warnings if issue.severity == "warning" else self.erros
        bucket.append(issue.to_dict())

    def to_dict(self) -> Dict[str, Any]:
        return {
            "modo": self.modo,
            "arquivo": self.arquivo,
            "linhas_lidas": self.linhas_lidas,
            "escolas_novas": self.escolas_novas,
            "escolas_atualizadas": self.escolas_atualizadas,
            "coordenadores_novos": self.coordenadores_novos,
            "coordenadores_atualizados": self.coordenadores_atualizados,
            "professores_novos": self.professores_novos,
            "professores_atualizados": self.professores_atualizados,
            "turmas_novas": self.turmas_novas,
            "turmas_atualizadas": self.turmas_atualizadas,
            "warnings": self.warnings,
            "erros": self.erros,
        }


# ──────────────────────────────────────────────────────────────────────
# Parser — XLSX e CSV
# ──────────────────────────────────────────────────────────────────────

def parse_planilha(path: Path) -> Tuple[List[Dict[str, Any]], List[Issue]]:
    """Lê arquivo, retorna (rows, issues_estruturais).

    Issues estruturais (header faltando, número errado de colunas)
    bloqueiam o pipeline — caller checa antes de prosseguir.
    """
    issues: List[Issue] = []
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        rows = _parse_xlsx(path)
    elif suffix == ".csv":
        rows = _parse_csv(path)
    else:
        issues.append(Issue(
            line=None, field=None, code="formato_nao_suportado",
            message=f"Formato {suffix} não suportado. Use .xlsx ou .csv.",
        ))
        return [], issues

    if not rows:
        issues.append(Issue(line=None, field=None, code="planilha_vazia",
                            message="Planilha sem dados."))
        return [], issues

    # Valida headers da 1ª linha
    headers_raw = list(rows[0].keys())
    headers = [h.strip().lower() if isinstance(h, str) else "" for h in headers_raw]
    headers_set = set(headers)
    missing = [h for h in REQUIRED_HEADERS if h not in headers_set]
    extras = [h for h in headers if h not in REQUIRED_HEADERS
              and h not in OPTIONAL_HEADERS]
    if missing:
        issues.append(Issue(
            line=None, field=None, code="headers_invalidos",
            message=(f"Headers obrigatórios faltando: {missing}. "
                     f"Got: {headers_raw}"),
        ))
        return [], issues
    if extras:
        # Não-bloqueante — só avisa.
        issues.append(Issue(
            line=None, field=None, code="headers_extras",
            message=(f"Colunas extras ignoradas: {extras}. "
                     f"Aceitas: {list(REQUIRED_HEADERS + OPTIONAL_HEADERS)}"),
            severity="warning",
        ))

    # Renormaliza chaves dos dicts pra lowercase
    normalized = []
    for r in rows:
        nr = {}
        for k, v in r.items():
            if not isinstance(k, str):
                continue
            nr[k.strip().lower()] = (v.strip() if isinstance(v, str) else v)
        normalized.append(nr)
    return normalized, issues


def _parse_xlsx(path: Path) -> List[Dict[str, Any]]:
    """openpyxl, primeira aba. Retorna list[dict] keyed por header."""
    from openpyxl import load_workbook
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for raw in rows[1:]:
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in raw):
            continue  # pula linhas vazias
        out.append({headers[i]: raw[i] for i in range(min(len(headers), len(raw)))})
    return out


def _parse_csv(path: Path) -> List[Dict[str, Any]]:
    out = []
    with path.open(encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for r in reader:
            if all((not v) or (isinstance(v, str) and not v.strip())
                   for v in r.values()):
                continue
            out.append(dict(r))
    return out


# ──────────────────────────────────────────────────────────────────────
# Validação
# ──────────────────────────────────────────────────────────────────────

def _validate_email(email: str) -> Tuple[bool, str]:
    try:
        result = validate_email(email, check_deliverability=False)
        return True, result.normalized
    except EmailNotValidError as exc:
        return False, str(exc)


def validate_rows(
    rows: List[Dict[str, Any]],
    *,
    escola_id_regex: Optional[str] = None,
    ano_letivo: Optional[int] = None,
) -> List[Issue]:
    """Valida cada linha + coerência global. Retorna lista de Issues.
    Errors bloqueiam commit; warnings são informativos."""
    issues: List[Issue] = []
    escola_id_regex_compiled = re.compile(
        escola_id_regex or os.getenv("PORTAL_ESCOLA_ID_REGEX",
                                      DEFAULT_ESCOLA_ID_REGEX)
    )

    # Estado pra checagens globais (acumulado linha a linha)
    escola_nome_by_id: Dict[str, str] = {}
    coord_nome_by_email: Dict[str, str] = {}
    coord_escola_by_email: Dict[str, str] = {}
    prof_nome_by_email: Dict[str, str] = {}
    prof_escolas_by_email: Dict[str, set] = {}
    turma_keys: set = set()  # (escola_id, turma_codigo, ano_letivo)

    # Determinar ano letivo (default = ano corrente)
    ano = ano_letivo or datetime.now(timezone.utc).year

    for idx, row in enumerate(rows, start=1):
        # ── Validação por campo ──
        # Só os obrigatórios bloqueiam — opcionais podem estar vazios.
        for field_name in REQUIRED_HEADERS:
            val = row.get(field_name)
            if val is None or (isinstance(val, str) and not val.strip()):
                issues.append(Issue(
                    line=idx, field=field_name, code="campo_vazio",
                    message=f"Campo obrigatório vazio: {field_name}",
                ))

        escola_id = row.get("escola_id") or ""
        escola_nome = row.get("escola_nome") or ""
        coord_email = row.get("coordenador_email") or ""
        coord_nome = row.get("coordenador_nome") or ""
        prof_email = row.get("professor_email") or ""
        prof_nome = row.get("professor_nome") or ""
        turma_codigo = row.get("turma_codigo") or ""
        turma_serie = row.get("turma_serie") or ""

        if escola_id and not escola_id_regex_compiled.match(escola_id):
            issues.append(Issue(
                line=idx, field="escola_id", code="escola_id_formato",
                message=(f"escola_id={escola_id!r} não casa com regex "
                         f"{escola_id_regex_compiled.pattern}"),
            ))

        if turma_codigo and not TURMA_CODIGO_REGEX.match(str(turma_codigo)):
            issues.append(Issue(
                line=idx, field="turma_codigo", code="turma_codigo_formato",
                message=(f"turma_codigo={turma_codigo!r} deve ser "
                         f"alfanumérico ≤8 chars"),
            ))

        if turma_serie and turma_serie not in VALID_SERIES:
            issues.append(Issue(
                line=idx, field="turma_serie", code="serie_invalida",
                message=(f"turma_serie={turma_serie!r} deve ser uma de "
                         f"{VALID_SERIES}"),
            ))

        if coord_email:
            ok, normalized = _validate_email(coord_email)
            if not ok:
                issues.append(Issue(
                    line=idx, field="coordenador_email",
                    code="email_invalido",
                    message=f"coordenador_email inválido: {normalized}",
                ))
            else:
                coord_email = normalized.lower()
                row["coordenador_email"] = coord_email

        if prof_email:
            ok, normalized = _validate_email(prof_email)
            if not ok:
                issues.append(Issue(
                    line=idx, field="professor_email",
                    code="email_invalido",
                    message=f"professor_email inválido: {normalized}",
                ))
            else:
                prof_email = normalized.lower()
                row["professor_email"] = prof_email

        # ── Coerência global ──
        if escola_id:
            if escola_id in escola_nome_by_id:
                if escola_nome_by_id[escola_id] != escola_nome:
                    issues.append(Issue(
                        line=idx, field="escola_nome",
                        code="escola_nome_inconsistente",
                        message=(f"escola_id={escola_id} tem nome "
                                 f"{escola_nome_by_id[escola_id]!r} em outra "
                                 f"linha mas {escola_nome!r} aqui"),
                    ))
            else:
                escola_nome_by_id[escola_id] = escola_nome

        if coord_email:
            if coord_email in coord_nome_by_email:
                if coord_nome_by_email[coord_email] != coord_nome:
                    issues.append(Issue(
                        line=idx, field="coordenador_nome",
                        code="coord_nome_inconsistente",
                        message=(f"coordenador_email={coord_email} tem nome "
                                 f"{coord_nome_by_email[coord_email]!r} em "
                                 f"outra linha mas {coord_nome!r} aqui"),
                    ))
                if (escola_id and coord_escola_by_email.get(coord_email)
                        and coord_escola_by_email[coord_email] != escola_id):
                    issues.append(Issue(
                        line=idx, field="coordenador_email",
                        code="coord_em_multiplas_escolas",
                        message=(f"coordenador_email={coord_email} aparece "
                                 f"em escolas diferentes "
                                 f"({coord_escola_by_email[coord_email]} e "
                                 f"{escola_id})"),
                    ))
            else:
                coord_nome_by_email[coord_email] = coord_nome
                if escola_id:
                    coord_escola_by_email[coord_email] = escola_id

        if prof_email:
            if prof_email in prof_nome_by_email:
                if prof_nome_by_email[prof_email] != prof_nome:
                    issues.append(Issue(
                        line=idx, field="professor_nome",
                        code="prof_nome_inconsistente",
                        message=(f"professor_email={prof_email} tem nome "
                                 f"{prof_nome_by_email[prof_email]!r} em "
                                 f"outra linha mas {prof_nome!r} aqui"),
                    ))
            else:
                prof_nome_by_email[prof_email] = prof_nome
            if escola_id:
                prof_escolas_by_email.setdefault(prof_email, set()).add(escola_id)

        if escola_id and turma_codigo:
            key = (escola_id, str(turma_codigo).upper(), ano)
            if key in turma_keys:
                issues.append(Issue(
                    line=idx, field="turma_codigo",
                    code="turma_duplicada",
                    message=(f"turma {turma_codigo} já apareceu em "
                             f"{escola_id}/{ano} em linha anterior"),
                ))
            else:
                turma_keys.add(key)

    # Warnings globais (não bloqueiam): professor em múltiplas escolas
    for email, escolas in prof_escolas_by_email.items():
        if len(escolas) > 1:
            issues.append(Issue(
                line=None, field="professor_email",
                code="prof_em_multiplas_escolas",
                message=(f"professor_email={email} ensina em "
                         f"{len(escolas)} escolas: {sorted(escolas)}. "
                         f"Caso legítimo, prosseguindo."),
                severity="warning",
            ))

    return issues


# ──────────────────────────────────────────────────────────────────────
# Geração de codigo_join
# ──────────────────────────────────────────────────────────────────────

def _slug_escola(codigo: str) -> str:
    """Últimos 5 chars do codigo, sem hífens. Ex.: SEDUC-CE-001 → CE001."""
    no_hyphen = codigo.replace("-", "")
    return no_hyphen[-5:].upper()


def gerar_codigo_join(
    session: Session, escola_codigo: str, turma_codigo: str, ano_letivo: int,
) -> str:
    """Gera codigo_join único. Em caso de colisão (raro), tenta sufixo
    -2, -3, ..."""
    base = f"TURMA-{_slug_escola(escola_codigo)}-{turma_codigo.upper()}-{ano_letivo}"
    if session.execute(
        select(Turma.id).where(Turma.codigo_join == base)
    ).first() is None:
        return base
    n = 2
    while True:
        candidate = f"{base}-{n}"
        if session.execute(
            select(Turma.id).where(Turma.codigo_join == candidate)
        ).first() is None:
            return candidate
        n += 1
        if n > 100:
            raise RuntimeError(f"Não consegui gerar codigo_join único pra {base}")


# ──────────────────────────────────────────────────────────────────────
# Sync ORM (idempotente)
# ──────────────────────────────────────────────────────────────────────

def sync_planilha(
    session: Session, rows: List[Dict[str, Any]], report: ImportReport,
    *, ano_letivo: Optional[int] = None,
) -> None:
    """Aplica upsert de Escola/Coord/Professor/Turma a partir das rows
    validadas. Atualiza contadores no `report`. Caller decide se commita
    ou rollback baseado em report.erros."""
    ano = ano_letivo or datetime.now(timezone.utc).year

    # Caches dentro do batch pra evitar flush por linha
    escolas_cache: Dict[str, Escola] = {}
    coords_cache: Dict[str, Coordenador] = {}
    profs_cache: Dict[str, Professor] = {}

    for idx, row in enumerate(rows, start=1):
        try:
            escola = _upsert_escola(session, row, escolas_cache, report)
            if escola is None:
                continue
            _upsert_coordenador(session, row, escola, coords_cache, report)
            professor = _upsert_professor(session, row, escola, profs_cache, report)
            if professor is None:
                continue
            _upsert_turma(session, row, escola, professor, ano, report)
        except Exception as exc:  # noqa: BLE001
            report.add_issue(Issue(
                line=idx, field=None, code="erro_sync",
                message=f"{type(exc).__name__}: {exc}",
            ))


def _upsert_escola(
    session: Session, row: Dict[str, Any], cache: Dict[str, Escola],
    report: ImportReport,
) -> Optional[Escola]:
    codigo = row.get("escola_id") or ""
    if not codigo:
        return None
    if codigo in cache:
        return cache[codigo]
    nome = row.get("escola_nome") or ""

    # M3: aceita escola_estado/escola_municipio explícitos da planilha.
    # Fallback: inferência do código.
    estado_explicito = (row.get("escola_estado") or "").strip().upper() or None
    municipio_explicito = (row.get("escola_municipio") or "").strip() or None
    if estado_explicito and len(estado_explicito) == 2 and estado_explicito.isalpha():
        estado = estado_explicito
    else:
        inferido = _extract_estado_from_codigo(codigo)
        if inferido is None:
            report.add_issue(Issue(
                line=None, field="escola_estado",
                code="estado_inferido_fallback",
                message=(f"Não consegui inferir UF de escola_id={codigo!r}. "
                         f"Usando fallback 'BR'. Adicione coluna "
                         f"`escola_estado` na planilha pra precisão."),
                severity="warning",
            ))
            estado = "BR"
        else:
            estado = inferido
    municipio = municipio_explicito or "(não informado)"

    existing = session.execute(
        select(Escola).where(Escola.codigo == codigo)
    ).scalar_one_or_none()
    if existing is None:
        escola = Escola(codigo=codigo, nome=nome, estado=estado,
                        municipio=municipio)
        session.add(escola)
        session.flush()
        report.escolas_novas += 1
    else:
        escola = existing
        changed = False
        if existing.nome != nome:
            existing.nome = nome
            changed = True
        if estado_explicito and existing.estado != estado:
            existing.estado = estado
            changed = True
        if municipio_explicito and existing.municipio != municipio:
            existing.municipio = municipio
            changed = True
        if changed:
            report.escolas_atualizadas += 1
    cache[codigo] = escola
    return escola


def _extract_estado_from_codigo(codigo: str) -> Optional[str]:
    """Tenta extrair UF do código no formato SEDUC-CE-001."""
    parts = codigo.split("-")
    for p in parts:
        if len(p) == 2 and p.isalpha() and p.isupper():
            return p
    return None


def _upsert_coordenador(
    session: Session, row: Dict[str, Any], escola: Escola,
    cache: Dict[str, Coordenador], report: ImportReport,
) -> Optional[Coordenador]:
    email = (row.get("coordenador_email") or "").lower()
    if not email:
        return None
    if email in cache:
        return cache[email]
    nome = row.get("coordenador_nome") or ""

    existing = session.execute(
        select(Coordenador).where(Coordenador.email == email)
    ).scalar_one_or_none()
    if existing is None:
        coord = Coordenador(escola_id=escola.id, nome=nome, email=email)
        session.add(coord)
        session.flush()
        report.coordenadores_novos += 1
    else:
        coord = existing
        changed = False
        if existing.escola_id != escola.id:
            existing.escola_id = escola.id
            changed = True
        if existing.nome != nome:
            existing.nome = nome
            changed = True
        if changed:
            report.coordenadores_atualizados += 1
    cache[email] = coord
    return coord


def _upsert_professor(
    session: Session, row: Dict[str, Any], escola: Escola,
    cache: Dict[str, Professor], report: ImportReport,
) -> Optional[Professor]:
    email = (row.get("professor_email") or "").lower()
    if not email:
        return None
    if email in cache:
        return cache[email]
    nome = row.get("professor_nome") or ""

    existing = session.execute(
        select(Professor).where(Professor.email == email)
    ).scalar_one_or_none()
    if existing is None:
        prof = Professor(escola_id=escola.id, nome=nome, email=email)
        session.add(prof)
        session.flush()
        report.professores_novos += 1
    else:
        prof = existing
        changed = False
        if existing.escola_id != escola.id:
            existing.escola_id = escola.id
            changed = True
        if existing.nome != nome:
            existing.nome = nome
            changed = True
        if changed:
            report.professores_atualizados += 1
    cache[email] = prof
    return prof


def _upsert_turma(
    session: Session, row: Dict[str, Any], escola: Escola,
    professor: Professor, ano_letivo: int, report: ImportReport,
) -> Optional[Turma]:
    codigo = (row.get("turma_codigo") or "").upper()
    serie = row.get("turma_serie") or ""
    if not codigo or not serie:
        return None

    existing = session.execute(
        select(Turma).where(
            Turma.escola_id == escola.id,
            Turma.codigo == codigo,
            Turma.ano_letivo == ano_letivo,
        )
    ).scalar_one_or_none()

    if existing is None:
        codigo_join = gerar_codigo_join(session, escola.codigo, codigo, ano_letivo)
        turma = Turma(
            escola_id=escola.id, professor_id=professor.id,
            codigo=codigo, serie=serie, codigo_join=codigo_join,
            ano_letivo=ano_letivo,
        )
        session.add(turma)
        session.flush()
        report.turmas_novas += 1
        return turma
    else:
        changed = False
        if existing.professor_id != professor.id:
            existing.professor_id = professor.id
            changed = True
        if existing.serie != serie:
            existing.serie = serie
            changed = True
        if changed:
            report.turmas_atualizadas += 1
        return existing


# ──────────────────────────────────────────────────────────────────────
# Pipeline orquestrador
# ──────────────────────────────────────────────────────────────────────

def run_import(
    session: Session, file_path: Path, *, modo: str = "dry-run",
    ano_letivo: Optional[int] = None,
    rollback_on_error: bool = True,
) -> ImportReport:
    """Pipeline completo: parse → validate → sync → relatório.

    `modo`:
    - "dry-run": NÃO chama session.commit. Caller deve rollback.
    - "commit": chama session.commit no final, OU rollback se
      rollback_on_error=True e há erros.
    """
    if modo not in ("dry-run", "commit"):
        raise ValueError(f"modo inválido: {modo}")

    report = ImportReport(modo=modo, arquivo=str(file_path))

    rows, parse_issues = parse_planilha(file_path)
    for issue in parse_issues:
        report.add_issue(issue)
    if parse_issues and any(i.severity == "error" for i in parse_issues):
        return report

    report.linhas_lidas = len(rows)
    val_issues = validate_rows(rows, ano_letivo=ano_letivo)
    for issue in val_issues:
        report.add_issue(issue)

    has_errors = any(i.get("severity") == "error" for i in report.erros)
    if has_errors and rollback_on_error and modo == "commit":
        return report

    sync_planilha(session, rows, report, ano_letivo=ano_letivo)

    has_errors_after = any(i.get("severity") == "error" for i in report.erros)
    if modo == "commit":
        if has_errors_after and rollback_on_error:
            session.rollback()
        else:
            session.commit()
    else:
        # dry-run: sempre rollback no final
        session.rollback()

    return report
